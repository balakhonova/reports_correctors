#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#этот скрипт нужно скопировать туда, где лежат скаченные (неисправленные) отчеты

# Для установки openpyxl:
# sudo python3 -m pip install openpyxl

import fnmatch
import openpyxl
import os
import shutil
import sys
from openpyxl.utils import get_column_letter

#все исправленные отчеты окажутся в папке renamed_xls, которая лежит в той же директории, что и данный скрипт
DIRNAME = "renamed_xls"

if not os.path.exists(DIRNAME):
    os.makedirs(DIRNAME)
 
#ввод команды в shell: ./corrector.py месяц
text = ""
if len(sys.argv) == 2:
    text = sys.argv[1]

#выношу цены в отдельный словарь для удобства возможных корректировок тарифов

correct_price = {"Промо":150, "Базовый":250, "Супербазовый":350, "Дождь":240, "МАТЧ! Футбол":380, "Настрой кино":380, "25 за 25":25,
                 "Ночной":199, "Первый Бандл":50,"Промо Бандл":100,"Базовый Бандл":250, "Супербазовый Бандл":350, "Amedia Premium HD":199,
                 "Наш футбол HD":219, "Наш футбол":219, "Trial":50, "SHANT Premium":240, "Публичный":990, "Поддержка":0, "Архив Бесплатный":0,
                 "Старт":0, "Архив Ночной":150, "Поддержка VOD":0,"Супербазовый Бандл 1/3":0, "Промо Бандл 1/3":0, "Базовый Бандл 1/3":0}
        
#Пройдусь по всем файликам, имеющим расширение xls, в текущей директории и заменю расширение на xlsx

for file in os.listdir('.'):
    if fnmatch.fnmatch(file, '*.xls'):
        print(file, end=': ')
        shutil.copy(file, file+'x')
        wb = openpyxl.load_workbook(file+'x')
        ws = wb.active
        company_name = ws.cell(row=1, column=1).value
        print(company_name)
        try:
            newfilename = DIRNAME + '/' + "Отчет " + text + " " + company_name + '.xlsx'
            slist = list(newfilename)
            for i, c in enumerate(slist):
                if slist[i] == 'ё':
                    slist[i] = 'е'
                    newfilename = ''.join(slist)
                if slist[i] == 'й':
                    slist[i] = 'и'
                    newfilename = ''.join(slist)
                if slist[i] == 'Ё':
                    slist[i] = 'Е'
                    newfilename = ''.join(slist)
                if slist[i] == 'Й':
                    slist[i] = 'И'
                    newfilename = ''.join(slist)
                if slist[i] == '«' or slist[i] == '»' or slist[i] == '"':
                    slist[i] = ''
                    newfilename = ''.join(slist)
        except TypeError:
            newfilename = DIRNAME + '/' + "Отчет " + text + " Undefined!!!!!!!!" + '.xlsx'
            print("В отчете", file, "не указано юр.лицо. Поправьте в CMS")
            
        # Если у какого-то пакета цена = 0, сделаю соответствующую замену
        
        for i in range(1,ws.max_column):
            if ws.cell(row=6,column=i).value == "Стоимость подписки":
                column_price_index = i
                col_reg = get_column_letter(i-1)
                wdt = ws.column_dimensions[col_reg].width = 16
        for price_cell in range(6, ws.max_row):
            current_price = ws.cell(row=price_cell ,column=column_price_index).value
            tariff = ws.cell(row=price_cell ,column=column_price_index - 1).value
            if current_price == 0 and not (tariff == "Поддержка" or tariff == "Старт" or tariff == "Архив Бесплатный" or tariff == "Лайт МА"):
                new_price = correct_price.get(tariff)
                ws.cell(row=price_cell ,column=column_price_index).value = new_price
                print("В строке ", ws.cell(row=price_cell,column=1).value, " отчета заменена цена")
        
        #Расширю столбец "Дата регистрации", чтобы он не был ########
        for i in range(1,ws.max_column):
            if ws.cell(row=6,column=i).value == "Дата регистрации":
                col_reg = get_column_letter(i)
                wdt = ws.column_dimensions[col_reg].width = 16
                
        # Передвину суммы
        for i in range(1,ws.max_column):
            if ws.cell(row=6,column=i).value == "Начисленная абонентская плата":
                #a = ws.cell(row=7,column=i).coordinate
                #b = ws.cell(row=ws.max_row-7,column=i).coordinate
                ws.cell(row=ws.max_row-5,column=i).value = "=SUM("+ ws.cell(row=7,column=i).coordinate +":"+ ws.cell(row=ws.max_row-7,column=i).coordinate +")"
                ws.cell(row=ws.max_row-5,column=i+1).value = "=SUM("+ ws.cell(row=7,column=i+1).coordinate +":"+ ws.cell(row=ws.max_row-7,column=i+1).coordinate +")"
                for j in range (1,7):
                    ws.cell(row=ws.max_row-5,column=i-j).value = ""

        #Проверка длительности
        for i in range(1,ws.max_column):
            if ws.cell(row=6,column=i).value == "Длительность предоставления услуги за отчетный период":
                for j in range(7,ws.max_row-7):
                    if ws.cell(row=j,column=i).value < 0:
                        print ("ОШИБКА В ДЛИТЕЛЬНОСТИ! В строке ",ws.cell(row=j,column=1).value," отчета")

        for i in range(1,ws.max_column):
            if ws.cell(row=6,column=i).value == "Идентификатор подписки":
                j = get_column_letter(i)
                ws.column_dimensions[j].hidden = True

                    
        # Передвину место для подписи
        ws.cell(row=1, column=10).value = "ОТЧЕТ АГЕНТА"    
        cellLS____ = ws.cell(row=ws.max_row-1, column=6)
        cellMPLS = ws.cell(row=ws.max_row, column=6)
        GenD = ws.cell(row=ws.max_row-1, column=1)
        ws.cell(row=ws.max_row , column=9).value = cellLS____.value
        ws.cell(row=ws.max_row +1, column=9).value = cellMPLS.value
        ws.cell(row=ws.max_row -1, column=1).value = GenD.value
        cellLS____.value = ''
        cellMPLS.value = ''
        GenD.value = ''

        wb.save(newfilename)

print ("ОБРАБОТКА ФАЙЛОВ ЗАВЕРШЕНА")

