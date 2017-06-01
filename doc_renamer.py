#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Для установки xlrd:
# sudo python3 -m pip install xlrd

import fnmatch
import openpyxl
import os
import shutil
import sys


DIRNAME = "renamed_xls"

if not os.path.exists(DIRNAME):
    os.makedirs(DIRNAME)

text = ""
if len(sys.argv) == 2:
	text = sys.argv[1]


for file in os.listdir('.'):
	if fnmatch.fnmatch(file, '*.xls'):
		print(file, end=': ')
		# Добавим х на конце, чтобы получилось xlsx
		shutil.copy(file, file+'x')
		wb = openpyxl.load_workbook(file+'x')
		ws = wb.active
		company_name =  ws.cell(row=1, column=1).value

		# Определим имя компаннии из 1-й ячейки
		print(company_name)
		# Новое имя файла:
		newfilename = DIRNAME + '/' + "Отчёт " + text + " " + company_name + '.xlsx'
		# print(newfilename)

		#shutil.copy(file, newfilename)

		#сделаем строку шире
		#ws.cell(row=ws.max_row-4, column=1).value = '\r\n\r\n'
		#ws.cell(row=ws.max_row-4, column=1).font = openpyxl.styles.Font(size=24, italic=True)

		#переместим место для подписей
		cellGend = ws.cell(row=ws.max_row-1, column=1)
		cell____ = ws.cell(row=ws.max_row-1, column=6)
		cellMP = ws.cell(row=ws.max_row, column=6)
		newColumn = 9
		ws.cell(row=ws.max_row - 3, column=newColumn).value = cellGend.value
		ws.cell(row=ws.max_row - 3, column=newColumn + 6).value = cell____.value
		ws.cell(row=ws.max_row - 2, column=newColumn + 6).value = cellMP.value
		cellGend.value = ''
		cell____.value = ''
		cellMP.value = ''

		wb.save(newfilename)



