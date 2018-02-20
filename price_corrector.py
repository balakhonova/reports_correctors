#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#этот скрипт нужно скопировать туда, где лежат ИСПРАВЛЕННЫЕ отчеты (папка renamed_xls), сюда же нужно скачать табличку Operators_tariffs, предварительно сняв закрепления


# Для установки openpyxl:
# sudo python3 -m pip install openpyxl

import fnmatch
import openpyxl
import os
import shutil
import sys
from openpyxl.utils import get_column_letter

#все исправленные отчеты окажутся в папке renamed_xls, которая лежит в той же директории, что и данный скрипт

db = openpyxl.load_workbook("OperatorsTariffs.xlsx")
dbs = db.active
Operators_list = []
for col in dbs['A']:
     Operators_list.append(col.value)
all_the_tariffs = []
for col in range(1, dbs.max_column):
    all_the_tariffs.append(dbs.cell(row=1, column=col).value)
log = open("log.txt", "w")

#Пройдусь по всем файликам, имеющим расширение xls, в текущей директории и заменю расширение на xlsx
for file in os.listdir('.'):
    if fnmatch.fnmatch(file, '*.xlsx') and not file == "OperatorsTariffs.xlsx":
        print(file, end=': \n')
        log.write('\n\n\n'+file+': \n')
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        op_name = ws.cell(row=1,column=1).value
        if not op_name in Operators_list:
            print("Оператора "+ op_name+ " нет в таблице \n")
            log.write("Оператора "+ op_name+ " нет в таблице \n")
        else:
            
            for i in range(1,ws.max_column):
                if ws.cell(row=6,column=i).value == "Стоимость подписки":
                    column_price_index = i
            
            for price_cell in range(7, ws.max_row-7):
                current_price = ws.cell(row=price_cell, column=column_price_index).value
                current_ag = ws.cell(row=price_cell, column=column_price_index + 1).value
                tariff = ws.cell(row=price_cell, column=column_price_index - 1).value
                if not tariff in all_the_tariffs:
                    print("Тарифа "+tariff+" нет в таблице \n")
                    log.write("Тарифа "+tariff+" нет в таблице \n")
                tariff_agent = "АВ "+ tariff
    
                for k in range(1, dbs.max_row):
                    if op_name == dbs.cell(row=k, column=1).value:
                        for n in range (3, dbs.max_column):
                            if n%2==0:
                                ag_name = dbs.cell(row=1,column=n).value
                                ag_list = dbs.cell(row=k,column=n).value
                                tariff_ag = {ag_name:ag_list}
                                new_ag = tariff_ag.get(tariff_agent)
                                if not new_ag == current_ag:
                                    if not new_ag == None:
                                        log.write("В строке "+ str(ws.cell(row=price_cell,column=1).value)+ " отчета заменено агентское вознаграждение за пакет "+tariff+ ": \n   Было: "+ str(current_ag)+ "; Стало: "+ str(new_ag) +"\n")
                                        ws.cell(row=price_cell ,column=column_price_index+1).value = new_ag
                            else:
                                tar_name = dbs.cell(row=1,column=n).value
                                tar_list = dbs.cell(row=k,column=n).value
                                tariff_price = {tar_name:tar_list}
                                new_price = tariff_price.get(tariff)
                                if not new_price == current_price:
                                    if not new_price == None:
                                        log.write(str("В строке "+ str(ws.cell(row=price_cell,column=1).value)+ " отчета заменена цена на пакет " +str(tariff)+ ": \n   Было: "+ str(current_price)+ "; Стало: "+ str(new_price)+ "\n"))
                                        ws.cell(row=price_cell ,column=column_price_index).value = new_price

    
            wb.save(file)
print ("ОБРАБОТКА ФАЙЛОВ ЗАВЕРШЕНА")
log.close()


